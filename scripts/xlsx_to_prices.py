# -*- coding: utf-8 -*-
"""
Parse quotation workbook(s) -> lib/catalog/xlsx-imported-products.ts

Supports:
  - 建议成交价 | 拟定成交价 | 含税*价 等价列名
  - 同一 sheet 内多套表（遇新表头则重置上下文）
  - 嵌入在工作表内的图片导出到 public/catalog/imported/

Regenerate examples:
  python scripts/xlsx_to_prices.py
  python scripts/xlsx_to_prices.py \"c:\\\\path\\\\簿1.xlsx\" \"c:\\\\path\\\\补充.xlsx\"
"""
from __future__ import annotations

import argparse
import hashlib
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "lib" / "catalog" / "xlsx-imported-products.ts"
IMPORT_ROOT = ROOT / "public" / "catalog" / "imported"
XLSX_DEFAULT = Path(r"c:\Users\21chy\Desktop\报价文件2025-7-22A1.xlsx")

# ProductSystemCategory
SYSTEM_MAP = [
    ("内门", "door", "内门", "Interior doors"),
    ("移门执手", "door", "移门执手", "Sliding door hardware"),
    ("移门", "door", "移门", "Sliding doors"),
    ("导轨", "hardware", "导轨", "Tracks & hardware"),
    ("墙板", "wall_panel", "墙板", "Wall panels"),
    ("墙", "wall_panel", "墙面", "Wall paneling"),
    ("整木", "cabinet", "整木", "Solid wood / millwork"),
    ("木作", "cabinet", "木作", "Millwork"),
    ("收纳", "storage", "收纳", "Storage"),
    ("浴室", "kitchen", "浴室", "Bathroom"),
    ("浴室柜", "kitchen", "浴室柜", "Bathroom vanity"),
    ("厨房", "kitchen", "厨房", "Kitchen"),
    ("灵动", "cabinet", "灵动", "Motiva"),
    ("照明", "lighting", "照明", "Lighting"),
    ("灯光", "lighting", "灯光", "Lighting"),
    ("五金", "hardware", "五金", "Hardware"),
]


def detect_system(sheet_name: str) -> tuple[str, str, str]:
    for key, sys, czh, cen in SYSTEM_MAP:
        if key in sheet_name:
            return sys, czh, cen
    return "cabinet", "综合", "General"


def norm_unit(u: str | None) -> str:
    if not u:
        return "pc"
    u = str(u).strip()
    if u in ("套",):
        return "set"
    if u in ("平方", "㎡", "平方米", "平米"):
        return "sqm"
    if u in ("米", "延米", "M", "m"):
        return "lm"
    if u in ("扇",):
        return "leaf"
    if u in ("件", "个", "支", "根"):
        return "pc"
    return "pc"


def norm_grade(g: str | None) -> str | None:
    if not g:
        return None
    g = str(g).strip().upper().replace("＋", "+")
    if "/" in g and len(g) < 10:
        return None
    if g in ("A+", "A", "B", "C", "D"):
        return g
    return None


def parse_interval(text: object) -> dict:
    """mm bounds + label — align with Tier matching in lib/pricing/engine.ts."""
    raw = "" if text is None else str(text).replace("\n", "; ").strip()
    d: dict = {"labelZh": raw, "labelEn": raw[:120] if raw else "Band"}
    if not raw or raw == "/":
        return d
    t = raw.replace(" ", "")

    m = re.search(
        r"[（(](\d+)\s*≤?\s*高\s*≤?\s*(\d+)\s*[）)]\s*[*×＊]\s*[（(](\d+)\s*≤?\s*宽\s*≤?\s*(\d+)\s*[）)]",
        t,
    )
    if m:
        d["heightMinMm"] = int(m.group(1))
        d["heightMaxMm"] = int(m.group(2))
        d["widthMinMm"] = int(m.group(3))
        d["widthMaxMm"] = int(m.group(4))
        return d

    m = re.search(
        r"宽\D*(\d+)\D*[*×x＊]\D*深\D*\(?(\d+)/(\d+)\)\D*[*×x＊]\D*高\D*[：:=為]?\D*(\d+)",
        t,
    )
    if m:
        wv = int(m.group(1))
        d["widthMinMm"] = wv
        d["widthMaxMm"] = wv
        d["depthMinMm"] = int(m.group(2))
        d["depthMaxMm"] = int(m.group(3))
        h = int(m.group(4))
        d["heightMinMm"] = h
        d["heightMaxMm"] = h
        return d

    m = re.search(
        r"(\d+)以上(\d+)\D*[*×x＊]\D*深\D*\(?(\d+)/(\d+)\)\D*[*×x＊]\D*高\D*[：:=為、]?\D*(\d+)",
        t,
    )
    if m:
        d["widthMinMm"] = int(m.group(1))
        d["widthMaxMm"] = int(m.group(2))
        d["depthMinMm"] = int(m.group(3))
        d["depthMaxMm"] = int(m.group(4))
        h = int(m.group(5))
        d["heightMinMm"] = h
        d["heightMaxMm"] = h
        return d

    m = re.search(
        r"(\d+)mm\D*[＜<]\D*(?:高度|高)\s*[＜≤]\s*(\d+)mm?",
        t,
    )
    if m:
        d["heightMinMm"] = int(m.group(1))
        d["heightMaxMm"] = int(m.group(2))
        return d

    m = re.search(r"高度\D*[≤＜＝=]\s*(\d+)mm?", t)
    if m:
        h = int(m.group(1))
        d["heightMaxMm"] = h
        if "heightMinMm" not in d:
            d["heightMinMm"] = h

    m = re.search(r"高\D*[≤＜≤]\s*(\d+)", t)
    if m and "heightMaxMm" not in d:
        d["heightMaxMm"] = int(m.group(1))
    m = re.search(r"宽\D*[≤＜≤]\s*(\d+)", t)
    if m:
        d["widthMaxMm"] = int(m.group(1))
    m = re.search(r"墙厚\D*[≤＜≤]\s*(\d+)", t)
    if m:
        d["depthMaxMm"] = int(m.group(1))

    m = re.search(
        r"（\s*(\d+)\s*≤?\s*高\s*≤?\s*(\d+)\s*）|[（(]\s*(\d+)\s*≤?\s*高\s*≤?\s*(\d+)\s*[）)]",
        t,
    )
    if m:
        grp = [g for g in m.groups() if g]
        if len(grp) >= 2:
            d["heightMinMm"] = int(grp[-2])
            d["heightMaxMm"] = int(grp[-1])
    m = re.search(
        r"（\s*(\d+)\s*≤?\s*宽\s*≤?\s*(\d+)\s*）|[（(]\s*(\d+)\s*≤?\s*宽\s*≤?\s*(\d+)\s*[）)]",
        t,
    )
    if m:
        grp = [g for g in m.groups() if g]
        if len(grp) >= 2:
            d["widthMinMm"] = int(grp[-2])
            d["widthMaxMm"] = int(grp[-1])

    sep = "*" in t or "×" in t or "＊" in t
    m = re.search(
        r"(\d+)\s*[*×x＊]\s*(\d+)\s*[*×x＊]\s*(\d+)(?:/(\d+))?",
        t,
    )
    if m and sep and "widthMinMm" not in d and "widthMaxMm" not in d:
        d["widthMinMm"] = int(m.group(1))
        d["widthMaxMm"] = int(m.group(1))
        d["depthMinMm"] = int(m.group(2))
        d["depthMaxMm"] = int(m.group(2))
        d["heightMinMm"] = int(m.group(3))
        d["heightMaxMm"] = int(m.group(3))
        if m.group(4):
            dh = int(m.group(4))
            if dh > int(m.group(3)):
                d["heightMaxMm"] = dh

    m = re.search(r"宽\D*(\d+)", t)
    if m and "widthMinMm" not in d:
        dv = int(m.group(1))
        d["widthMinMm"] = dv
        d["widthMaxMm"] = dv
    m = re.search(r"深\D*(\d+)", t)
    if m:
        dv = int(m.group(1))
        d.setdefault("depthMinMm", dv)
        d.setdefault("depthMaxMm", dv)

    return d


def is_price(x) -> bool:
    if x is None:
        return False
    if isinstance(x, (int, float)):
        if isinstance(x, float) and (x != x):
            return False
        return x > 0
    s = str(x).strip()
    if s in ("#REF!", "#VALUE!", "-", "") or s.endswith("!"):
        return False
    try:
        return float(s) > 0
    except ValueError:
        return False


def to_price(x) -> float:
    if isinstance(x, (int, float)):
        return float(x)
    return float(str(x).strip())


def build_column_map(cells: list[str]) -> dict[str, int] | None:
    cmap: dict[str, int] = {}
    for i, h in enumerate(cells):
        if not h:
            continue
        hl = str(h).strip()
        if "产品" in hl and ("名称" in hl or hl == "品名"):
            cmap["name"] = i
        elif hl.endswith("型号") or hl == "型号":
            cmap["model"] = i
        elif "规格" in hl or "门洞" in hl or "洞规" in hl or "订货尺寸" in hl:
            cmap["spec"] = i
        elif "型材" in hl:
            cmap["profile"] = i
        elif "饰面" in hl or "玻璃" in hl:
            cmap["finish"] = i
        elif "级别" in hl or hl == "等级":
            cmap["grade"] = i
        elif "单位" in hl:
            cmap["unit"] = i
        elif "性质" in hl:
            cmap["nature"] = i
        elif ("建议" in hl or "拟定" in hl or "含税" in hl) and (
            "成交" in hl or "价" in hl
        ):
            cmap["suggested"] = i
        elif "门店" in hl or ("统一" in hl and "价" in hl):
            cmap["retail"] = i
        elif "折扣" in hl and "系数" in hl:
            cmap["discount"] = i
        elif "备注" in hl:
            cmap["note"] = i
        elif "图片" in hl or hl == "图":
            cmap["image"] = i
    if "suggested" not in cmap:
        return None
    if "model" not in cmap and "name" not in cmap:
        return None
    return cmap


def sanitize_sheet_slug(name: str) -> str:
    """ASCII-only prefixes for filenames (stable paths for static hosting)."""
    digest = hashlib.sha1(name.encode("utf-8")).hexdigest()[:10]
    letters = "".join(c for c in name if ord(c) < 128 and c.isalnum())[:28]
    return f"{letters}_{digest}" if letters else f"x_{digest}"


def collect_embedded_images(
    ws, sheet_slug: str, seen_hashes: dict[str, int]
) -> dict[int, list[str]]:
    IMPORT_ROOT.mkdir(parents=True, exist_ok=True)
    out: dict[int, list[str]] = {}
    images = getattr(ws, "_images", None) or []

    def write_blob(data_bytes: bytes, row1: int, ix: int) -> None:
        if data_bytes[:4] == b"\xff\xd8\xff\xe0" or data_bytes[:2] == b"\xff\xd8":
            suf = ".jpg"
        elif len(data_bytes) > 12 and data_bytes[:4] == b"RIFF" and data_bytes[8:12] == b"WEBP":
            suf = ".webp"
        else:
            suf = ".png"
        stem = hashlib.sha1(data_bytes[:8192]).hexdigest()[:12]
        fname = f"{sheet_slug}_{row1}_{ix}_{stem}{suf}"
        cnt = seen_hashes.get(fname, 0)
        seen_hashes[fname] = cnt + 1
        if cnt:
            fname = f"{sheet_slug}_{row1}_{ix}_{stem}_{cnt}{suf}"
        path = IMPORT_ROOT / fname
        if not path.exists():
            path.write_bytes(data_bytes)
        out.setdefault(row1, []).append(f"/catalog/imported/{fname}")

    for ix, img in enumerate(images):
        anchor = getattr(img, "anchor", None)
        frm = getattr(anchor, "_from", None) if anchor else None
        row1 = int(frm.row) + 1 if frm else 0
        if row1 < 1:
            continue

        data_bytes = None
        try:
            if hasattr(img, "_data"):
                data_bytes = img._data()
        except Exception:
            data_bytes = None

        if data_bytes:
            write_blob(bytes(data_bytes), row1, ix)

    return out


def norm_image_ref(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ("/", "-"):
        return None
    if s.isdigit():
        return None
    if s.startswith("http://") or s.startswith("https://"):
        return s
    if s.startswith("/"):
        return s
    base = s.replace("\\", "/").split("/")[-1]
    if re.match(r"^[\w\-.]+\.(png|jpe?g|gif|webp|svg)$", base, re.I):
        return "/catalog/" + base
    return None


def uid(*parts: str) -> str:
    h = hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()[:10]
    return f"x-{h}"


def fmt_num(x: float) -> str:
    if isinstance(x, float) and x == int(x):
        return str(int(x))
    return str(x)


def esc(s: str) -> str:
    return (
        s.replace("\\", "\\\\")
        .replace('"', '\\"')
        .replace("\n", "\\n")
        .replace("\r", "")
    )


def ingest_workbook(path: Path, products: dict[str, dict], source_names: list[str]) -> None:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    source_names.append(path.name)

    seen_hashes: dict[str, int] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        slug = sanitize_sheet_slug(sheet_name)
        imgs_by_row = collect_embedded_images(ws, slug, seen_hashes)

        cmap: dict[str, int] | None = None
        prev_name = ""
        prev_model = ""
        prev_finish = ""
        prev_profile = ""

        sys, cat_zh, cat_en = detect_system(sheet_name)
        max_r = ws.max_row or 1
        for ri, row in enumerate(
            ws.iter_rows(min_row=1, max_row=max_r, values_only=True),
            start=1,
        ):
            if not row:
                continue
            cells = [str(c).strip() if c is not None else "" for c in row]
            if not any(cells):
                continue

            new_map = build_column_map(cells)
            if new_map is not None:
                cmap = new_map
                prev_name = ""
                prev_model = ""
                prev_finish = ""
                prev_profile = ""
                continue

            if cmap is None:
                continue

            def gv(key: str, default=""):
                if key not in cmap:
                    return default
                ii = cmap[key]
                if ii >= len(row):
                    return default
                vv = row[ii]
                return default if vv is None else vv

            name = gv("name")
            if name:
                prev_name = str(name).strip()
            model = gv("model")
            if model:
                prev_model = str(model).strip().replace("\n", " ")
            finish = gv("finish")
            if finish and str(finish).strip() not in ("/", ""):
                prev_finish = str(finish).strip().replace("\n", "; ")
            profile = gv("profile")
            if profile and str(profile).strip() not in ("/", ""):
                prev_profile = str(profile).strip()

            spec = gv("spec")
            grade = norm_grade(gv("grade"))
            unit = norm_unit(gv("unit"))
            sug = gv("suggested")
            ret = gv("retail") if "retail" in cmap else None
            note = gv("note")
            if not is_price(sug):
                continue
            if not prev_model and not prev_name:
                continue

            p_sug = to_price(sug)
            p_ret = to_price(ret) if is_price(ret) else None

            prod_key = f"{path.stem[:24]}|{sheet_name}|{prev_model}|{prev_name[:32]}"
            if prod_key not in products:
                pid = uid(path.name, sheet_name, prev_model, prev_name)
                products[prod_key] = {
                    "id": pid,
                    "sheet": sheet_name,
                    "system": sys,
                    "categoryZh": cat_zh,
                    "categoryEn": cat_en,
                    "productNameZh": prev_name or prev_model,
                    "productNameEn": prev_name or prev_model,
                    "model": prev_model or "—",
                    "defaultUnit": unit,
                    "exclusionZh": "",
                    "exclusionEn": "",
                    "notesZh": str(note).strip() if note else "",
                    "customRules": [],
                    "rows": [],
                }
            if img_u := (
                norm_image_ref(gv("image")) if "image" in cmap else None
            ):
                products[prod_key].setdefault("imageUrls", set()).add(img_u)

            for r_pix in (ri - 1, ri, ri + 1):
                for webp in imgs_by_row.get(r_pix, []):
                    products[prod_key].setdefault("imageUrls", set()).add(webp)

            iv = parse_interval(spec)
            interval_id = uid(prod_key, str(spec), str(grade), str(p_sug))
            if not grade:
                grade = "A"
            row_id = uid(interval_id, grade, str(p_sug))
            desc_zh = prev_finish or prev_profile or "—"
            products[prod_key]["rows"].append(
                {
                    "id": row_id,
                    "intervalId": interval_id,
                    "interval": iv,
                    "grade": grade,
                    "finishZh": desc_zh[:200],
                    "finishEn": desc_zh[:200],
                    "unit": unit,
                    "suggested": p_sug,
                    "retail": p_ret,
                }
            )

    wb.close()


def emit_typescript(products: dict[str, dict], source_label: str) -> None:
    safe_src = source_label.replace("*/", "")
    lines = [
        "/**",
        f" * Auto-generated from: {safe_src}",
        " * Excel 成交价列 -> rrpCnyPerUnit；门店价（如有） -> nationalRetailCnyPerUnit",
        " * Regenerate: `python scripts/xlsx_to_prices.py [workbook.xlsx ...]`",
        " */",
        "",
        'import type { PriceLibraryProduct, PriceLibraryRow } from "../domain/types";',
        "",
        "function R(",
        "  productId: string,",
        '  partial: Omit<PriceLibraryRow, "productId">,',
        "): PriceLibraryRow {",
        "  return { ...partial, productId };",
        "}",
        "",
        "export const XLSX_IMPORTED_PRODUCTS: PriceLibraryProduct[] = [",
    ]

    for _, p in sorted(products.items(), key=lambda x: (x[1]["model"], x[1]["productNameZh"])):
        if not p["rows"]:
            continue
        ex_zh = esc(p["notesZh"][:500]) if p["notesZh"] else ""
        lines.append("  {")
        lines.append(f'    id: "{esc(p["id"])}",')
        lines.append(f'    system: "{p["system"]}",')
        lines.append(f'    categoryZh: "{esc(p["categoryZh"])}",')
        lines.append(f'    categoryEn: "{esc(p["categoryEn"])}",')
        lines.append(f'    productNameZh: "{esc(p["productNameZh"])}",')
        lines.append(f'    productNameEn: "{esc(str(p["productNameEn"])[:80])}",')
        lines.append(f'    model: "{esc(p["model"])}",')
        lines.append(f'    defaultPricingUnit: "{p["defaultUnit"]}",')
        lines.append("    customRules: [],")
        lines.append(f'    exclusionZh: "{ex_zh}",')
        lines.append(f'    exclusionEn: "",')
        if p["notesZh"]:
            lines.append(f'    notesZh: "{ex_zh}",')
        iurls = sorted(p.get("imageUrls") or [])
        if iurls:
            urls_ts = ", ".join(f'"{esc(u)}"' for u in iurls)
            lines.append(f"    productImageUrls: [{urls_ts}],")
        lines.append("    rows: [")
        for r in p["rows"]:
            iv = r["interval"]
            lines.append(f'      R("{esc(p["id"])}", {{')
            lines.append(f'        id: "{esc(r["id"])}",')
            lines.append("        interval: {")
            lines.append(f'          id: "{esc(r["intervalId"])}",')
            lines.append(f'          labelZh: "{esc(iv.get("labelZh", ""))}",')
            lines.append(f'          labelEn: "{esc(iv.get("labelEn", ""))}",')
            for k in (
                "heightMinMm",
                "heightMaxMm",
                "widthMinMm",
                "widthMaxMm",
                "depthMinMm",
                "depthMaxMm",
            ):
                if k in iv:
                    lines.append(f"          {k}: {iv[k]},")
            lines.append("        },")
            lines.append(f'        finishGrade: "{r["grade"]}",')
            lines.append(f'        finishDescriptionZh: "{esc(r["finishZh"])}",')
            lines.append(f'        finishDescriptionEn: "{esc(r["finishEn"])}",')
            lines.append(f'        pricingUnit: "{r["unit"]}",')
            lines.append(f"        rrpCnyPerUnit: {fmt_num(r['suggested'])},")
            if r["retail"]:
                lines.append(
                    f"        nationalRetailCnyPerUnit: {fmt_num(r['retail'])},"
                )
            lines.append("      }),")
        lines.append("    ],")
        lines.append("  },")

    lines.append("];")
    lines.append("")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    ap = argparse.ArgumentParser(description="XLSX -> xlsx-imported-products.ts")
    ap.add_argument(
        "paths",
        nargs="*",
        help="Workbooks to merge (.xlsx). Empty = built-in desktop default.",
    )
    args = ap.parse_args()

    paths = [Path(p) for p in args.paths] if args.paths else [XLSX_DEFAULT]
    for p in paths:
        if not p.exists():
            print("Missing xlsx:", p, file=sys.stderr)
            sys.exit(1)

    products: dict[str, dict] = {}
    src_parts: list[str] = []
    for p in paths:
        ingest_workbook(p, products, src_parts)

    label = "; ".join(src_parts) if src_parts else "unknown"
    emit_typescript(products, label)

    print(f"Wrote {OUT} ({len(products)} products) from [{label}]")


if __name__ == "__main__":
    main()
