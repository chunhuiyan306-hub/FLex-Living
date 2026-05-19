"""Dump first rows of quotation xlsx — UTF-8 output to stdout (set PYTHONUTF8=1 on Windows).

Edit `path` to your workbook before running:

  PYTHONUTF8=1 python scripts/inspect_xlsx.py
"""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DUMP = ROOT / "_inspect_xlsx_dump.json"

path = r"c:\Users\21chy\Desktop\报价文件2025-7-22A1.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

out_obj: dict = {"path": path, "sheets": wb.sheetnames, "peek": {}}
for name in wb.sheetnames[:14]:
    ws = wb[name]
    rows = list(ws.iter_rows(max_row=12, values_only=True))
    out_obj["peek"][name] = [list(r) for r in rows]
wb.close()

DUMP.write_text(json.dumps(out_obj, ensure_ascii=False, indent=2), encoding="utf-8")
sys.stdout.buffer.write(DUMP.read_bytes())
